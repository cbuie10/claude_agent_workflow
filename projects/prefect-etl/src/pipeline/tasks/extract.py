"""Extract tasks — fetch raw data from external sources."""

from io import BytesIO

import httpx
from openpyxl import load_workbook
from prefect import task


@task(name="extract_earthquake_data", retries=2, retry_delay_seconds=10)
def extract_earthquake_data(api_url: str) -> dict:
    """Fetch earthquake GeoJSON from the USGS API."""
    response = httpx.get(api_url, timeout=30.0)
    response.raise_for_status()
    return response.json()


@task(name="extract_weather_data", retries=2, retry_delay_seconds=10)
def extract_weather_data(api_url: str) -> dict:
    """Fetch weather forecast JSON from the Open-Meteo API."""
    response = httpx.get(api_url, timeout=30.0)
    response.raise_for_status()
    return response.json()


@task(name="extract_occ_wells_data", retries=2, retry_delay_seconds=10)
def extract_occ_wells_data(csv_url: str) -> str:
    """Fetch Oklahoma Corporation Commission Wells CSV data.

    Returns raw CSV text as a string. The file is large (~126 MB),
    so timeout is set to 120 seconds.
    """
    response = httpx.get(csv_url, timeout=120.0)
    response.raise_for_status()
    return response.text


@task(name="extract_well_transfers", retries=2, retry_delay_seconds=10)
def extract_well_transfers(xlsx_url: str) -> list[tuple]:
    """Fetch Oklahoma Corporation Commission Well Transfers Excel data.

    Downloads the .xlsx file, parses it with openpyxl, and returns
    a list of row tuples (skipping the header row). The file is small
    (~943 rows), so timeout is set to 60 seconds.
    """
    response = httpx.get(xlsx_url, timeout=60.0)
    response.raise_for_status()

    # Parse Excel file from bytes
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    sheet = workbook.active

    # Convert to list of tuples, skipping header row
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:  # Skip header
            continue
        rows.append(row)

    return rows
